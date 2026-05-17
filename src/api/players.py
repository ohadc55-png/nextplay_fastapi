"""Players CRUD + metrics + bulk add.

Async port of `backend/api/players.py`. JSON `bulk` endpoint plus file-
based import (CSV / XLSX). pandas + openpyxl runs in `asyncio.to_thread`
so we don't block the event loop on a 200-row spreadsheet.
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from src.api.deps.auth import get_current_user
from src.core.database import get_db
from src.models.players import Player, PlayerMetric
from src.models.users import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["players"])


def _serialize_player(p: Player) -> dict:
    return {
        "id": p.id, "user_id": p.user_id, "team_id": p.team_id,
        "name": p.name, "number": p.number,
        "position": p.position or "", "height": p.height or "",
        "weight": p.weight or "", "age": p.age,
        "strengths": p.strengths or "", "weaknesses": p.weaknesses or "",
        "notes": p.notes or "",
        "dominant_hand": p.dominant_hand or "",
        "active": bool(p.active),
        "photo_url": p.photo_url or "",
        "scout_summary": p.scout_summary or "",
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }


def _require_team(user: User) -> int:
    if user.active_team_id is None:
        raise HTTPException(status_code=400, detail="no active team")
    return user.active_team_id


# ---------------------------------------------------------------------------
# Single player CRUD
# ---------------------------------------------------------------------------

class _PlayerCreateBody(BaseModel):
    name: str = Field(min_length=1, max_length=255)
    number: int | None = None
    position: str | None = ""
    height: str | None = ""
    weight: str | None = ""
    age: int | None = None
    strengths: str | None = ""
    weaknesses: str | None = ""
    notes: str | None = ""
    dominant_hand: str | None = ""


class _PlayerUpdateBody(BaseModel):
    name: str | None = None
    number: int | None = None
    position: str | None = None
    height: str | None = None
    weight: str | None = None
    age: int | None = None
    strengths: str | None = None
    weaknesses: str | None = None
    notes: str | None = None
    dominant_hand: str | None = None
    active: bool | None = None


@router.get("/players")
async def list_players(
    team_id: int | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Phase 15 — list the coach's roster as JSON.

    Used by the calendar's attendance modal to render the player list.
    Defaults to the active team when team_id is omitted. Returns
    `{"players": [...]}` shaped via `_serialize_player`.

    Cross-tenant safe: filter is `Player.user_id == coach.id`, so a
    coach can never see another coach's roster even by guessing a
    team_id.
    """
    from sqlalchemy import select

    tid = team_id if team_id is not None else user.active_team_id
    if tid is None:
        return {"players": []}
    stmt = (
        select(Player)
        .where(
            Player.user_id == user.id,
            Player.team_id == tid,
            Player.active.is_(True),
        )
        .order_by(Player.number.is_(None), Player.number, Player.name)
    )
    rows = list((await db.execute(stmt)).scalars().all())
    return {"players": [_serialize_player(p) for p in rows]}


@router.post("/player", status_code=201)
async def add_player(
    body: _PlayerCreateBody,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    team_id = _require_team(user)
    p = Player(
        user_id=user.id, team_id=team_id, name=body.name.strip(),
        number=body.number, position=body.position or "",
        height=body.height or "", weight=body.weight or "",
        age=body.age, strengths=body.strengths or "",
        weaknesses=body.weaknesses or "", notes=body.notes or "",
        dominant_hand=body.dominant_hand or "",
        active=True,
    )
    db.add(p)
    await db.flush()
    return _serialize_player(p)


@router.put("/player/{player_id}")
async def update_player(
    player_id: int,
    body: _PlayerUpdateBody,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    p = (await db.execute(
        select(Player).where(Player.id == player_id, Player.user_id == user.id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Player not found")

    data = body.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(p, key, value)
    await db.flush()
    return _serialize_player(p)


@router.delete("/player/{player_id}")
async def delete_player(
    player_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Soft delete: flip active=False (matches v1 — preserves history,
    stats, notebook references)."""
    existing = (await db.execute(
        select(Player.id).where(Player.id == player_id, Player.user_id == user.id)
    )).scalar_one_or_none()
    if not existing:
        raise HTTPException(status_code=404, detail="Player not found")
    await db.execute(
        update(Player).where(Player.id == player_id).values(active=False)
    )
    await db.flush()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Metrics
# ---------------------------------------------------------------------------

class _MetricsBody(BaseModel):
    metrics: dict = Field(default_factory=dict)


@router.post("/player/{player_id}/metrics")
async def save_player_metrics(
    player_id: int,
    body: _MetricsBody,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Upsert PlayerMetric (UNIQUE on player_id). Sets metrics_filled_at
    on the player row so the UI can show 'metrics complete' badges."""
    p = (await db.execute(
        select(Player).where(Player.id == player_id, Player.user_id == user.id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Player not found")

    row = (await db.execute(
        select(PlayerMetric).where(PlayerMetric.player_id == player_id)
    )).scalar_one_or_none()
    if row:
        row.metrics_json = body.metrics or {}
        row.updated_at = datetime.utcnow().isoformat()
    else:
        row = PlayerMetric(
            player_id=player_id, user_id=user.id, team_id=p.team_id,
            metrics_json=body.metrics or {},
            updated_at=datetime.utcnow().isoformat(),
        )
        db.add(row)

    p.metrics_filled_at = datetime.utcnow()
    await db.flush()
    return {"ok": True, "metrics": row.metrics_json}


# ---------------------------------------------------------------------------
# Bulk add (JSON + file uploads)
# ---------------------------------------------------------------------------

_ROSTER_TEMPLATE_COLUMNS = [
    "name", "number", "position", "height", "weight", "age",
    "strengths", "weaknesses", "notes",
]
_ROSTER_TEMPLATE_EXAMPLES = [
    ["John Smith", 7, "PG", "6'0\"", 180, 22,
     "Fast, good vision", "Weak left hand", "Team captain"],
    ["Mike Johnson", 23, "SF", "6'6\"", 210, 24,
     "Strong rebounder", "Limited range", ""],
]


class _BulkAddBody(BaseModel):
    players: list[dict] = Field(default_factory=list)


def _insert_player_rows(db: AsyncSession, user_id: int, team_id: int, players: list[dict]) -> tuple[int, int]:
    inserted = 0
    skipped = 0
    for raw in players:
        name = (raw.get("name") or "").strip()
        if not name:
            skipped += 1
            continue
        try:
            number = int(raw.get("number") or 0) or None
        except (TypeError, ValueError):
            number = None
        try:
            age = int(raw.get("age") or 0) or None
        except (TypeError, ValueError):
            age = None
        db.add(Player(
            user_id=user_id, team_id=team_id, name=name,
            number=number, position=(raw.get("position") or "").strip(),
            height=(raw.get("height") or "").strip(),
            weight=(raw.get("weight") or "").strip(),
            age=age,
            strengths=(raw.get("strengths") or "").strip(),
            weaknesses=(raw.get("weaknesses") or "").strip(),
            notes=(raw.get("notes") or "").strip(),
            active=True,
        ))
        inserted += 1
    return inserted, skipped


@router.post("/players/bulk")
async def bulk_add_players(
    body: _BulkAddBody,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Insert many players at once. Skips empty-name rows. Returns the
    number actually inserted so the UI can show 'imported X / skipped Y'."""
    team_id = _require_team(user)
    inserted, skipped = _insert_player_rows(db, user.id, team_id, body.players)
    await db.flush()
    return {"ok": True, "inserted": inserted, "skipped": skipped}


# ---------------------------------------------------------------------------
# Roster template downloads + file upload
# ---------------------------------------------------------------------------

@router.get("/players/template.csv")
async def roster_template_csv(_user: User = Depends(get_current_user)) -> Response:
    """Download a CSV template with the correct column headers + 2 example rows."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_ROSTER_TEMPLATE_COLUMNS)
    for row in _ROSTER_TEMPLATE_EXAMPLES:
        writer.writerow(row)
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=nextplay_roster_template.csv"},
    )


def _build_xlsx_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(_ROSTER_TEMPLATE_COLUMNS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in _ROSTER_TEMPLATE_EXAMPLES:
        ws.append(row)
    for col_idx, col_name in enumerate(_ROSTER_TEMPLATE_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(col_name) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/players/template.xlsx")
async def roster_template_xlsx(_user: User = Depends(get_current_user)) -> StreamingResponse:
    """Download an XLSX template with the correct column headers + 2 example rows."""
    data = await asyncio.to_thread(_build_xlsx_template)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nextplay_roster_template.xlsx"},
    )


def _parse_roster_file(filename: str, content: bytes) -> list[dict]:
    """Parse a CSV/XLSX roster file into a list of player dicts.
    Runs synchronously — wrap in asyncio.to_thread at the call site.
    """
    import pandas as pd

    name_lower = filename.lower()
    if name_lower.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(content), engine="openpyxl" if name_lower.endswith(".xlsx") else None)
    elif name_lower.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content))
    else:
        raise ValueError("Unsupported file type")
    df.columns = [str(c).strip().lower() for c in df.columns]

    def _s(row, key: str) -> str:
        v = row.get(key, "")
        return "" if pd.isna(v) else str(v).strip()

    def _i(row, key: str) -> int:
        v = row.get(key, 0)
        try:
            return int(float(v)) if not pd.isna(v) else 0
        except (ValueError, TypeError):
            return 0

    out: list[dict] = []
    for _, row in df.iterrows():
        name = _s(row, "name")
        if not name or name.lower() == "nan":
            continue
        out.append({
            "name": name,
            "number": _i(row, "number"),
            "position": _s(row, "position"),
            "height": _s(row, "height"),
            "weight": _s(row, "weight"),
            "age": _i(row, "age"),
            "strengths": _s(row, "strengths"),
            "weaknesses": _s(row, "weaknesses"),
            "notes": _s(row, "notes"),
        })
    return out


@router.post("/players/bulk-file")
async def bulk_add_players_file(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Accept a CSV or XLSX upload and bulk-import the rows as players."""
    team_id = _require_team(user)
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    content = await file.read()
    try:
        players = await asyncio.to_thread(_parse_roster_file, file.filename, content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from None
    except Exception as e:
        logger.warning("Failed to parse roster file: %s", e)
        raise HTTPException(status_code=400, detail="Could not read file. Please use the provided template.") from None

    inserted, skipped = _insert_player_rows(db, user.id, team_id, players)
    await db.flush()
    return {"ok": True, "count": inserted, "inserted": inserted, "skipped": skipped}
